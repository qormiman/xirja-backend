"""
Applies your reviewed decisions from review_medium_matches.xlsx (produced by
export_medium_matches.py / the "Export medium matches for review" workflow)
back to the database.

For each row on the 'Review' tab:

    decision = "keep"    -> marks the product as manually confirmed
                             (match_confidence = 'manual'). A manually
                             confirmed product is never automatically
                             downgraded by product_matcher.py.
    decision = "reject"  -> unlinks every listing currently attached to
                             that product and deletes the product row. The
                             listings go back to being unmatched and are
                             re-evaluated the next time product_matcher.py
                             runs.
    blank / anything else -> left alone. Still medium confidence, safe to
                              review again in a future export.

Writes are batched and retried on deadlock, exactly like product_matcher.py
-- this can safely run while a crawl or the matcher is also writing to the
database.

Run via GitHub Actions ("Apply reviewed matches" workflow) after uploading
your edited review_medium_matches.xlsx back into the repo root (same
filename, replacing the previous copy).
"""
import sys
import uuid

import psycopg2.extras
from openpyxl import load_workbook

from product_matcher import (
    DB_WRITE_HARD_TIMEOUT_SECONDS,
    WRITE_BATCH_SIZE,
    _chunks,
    _run_batch,
    get_connection,
    run_with_timeout,
)

INPUT_PATH = "review_medium_matches.xlsx"
SHEET_NAME = "Review"


def read_decisions(path):
    """Returns (keep_ids, reject_ids, skipped_count, bad_rows). bad_rows is
    [(raw_id, canonical_name, decision_text), ...] for anything that
    couldn't be understood, so it can be reported rather than silently
    dropped."""
    wb = load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Expected a '{SHEET_NAME}' sheet in {path}, found: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    keep_ids = []
    reject_ids = []
    skipped = 0
    bad_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue  # blank row, e.g. trailing rows Excel sometimes leaves behind

        raw_id = row[0]
        canonical_name = row[1] if len(row) > 1 else None
        raw_decision = row[4] if len(row) > 4 else None
        decision = (raw_decision or "").strip().lower()

        try:
            product_id = str(uuid.UUID(str(raw_id)))
        except (ValueError, AttributeError, TypeError):
            bad_rows.append((raw_id, canonical_name, raw_decision))
            continue

        if decision == "keep":
            keep_ids.append(product_id)
        elif decision == "reject":
            reject_ids.append(product_id)
        elif decision == "":
            skipped += 1
        else:
            bad_rows.append((raw_id, canonical_name, raw_decision))

    return keep_ids, reject_ids, skipped, bad_rows


def _bulk_confirm(cur, product_ids):
    psycopg2.extras.execute_values(
        cur,
        """
        UPDATE product AS p SET match_confidence = 'manual'
        FROM (VALUES %s) AS v(product_id)
        WHERE p.id = v.product_id::uuid
        """,
        [(pid,) for pid in product_ids],
    )
    return cur.rowcount


def _bulk_reject(cur, product_ids):
    psycopg2.extras.execute_values(
        cur,
        """
        UPDATE listing AS l SET product_id = NULL
        FROM (VALUES %s) AS v(product_id)
        WHERE l.product_id = v.product_id::uuid
        """,
        [(pid,) for pid in product_ids],
    )
    n_unlinked = cur.rowcount

    psycopg2.extras.execute_values(
        cur,
        """
        DELETE FROM product AS p
        USING (VALUES %s) AS v(product_id)
        WHERE p.id = v.product_id::uuid
        """,
        [(pid,) for pid in product_ids],
    )
    n_deleted = cur.rowcount

    return n_unlinked, n_deleted


def apply_all(conn, cur, keep_ids, reject_ids):
    confirmed = 0
    for chunk in _chunks(keep_ids, WRITE_BATCH_SIZE):
        confirmed += _run_batch(
            conn, lambda chunk=chunk: _bulk_confirm(cur, chunk), "confirming reviewed matches"
        ) or 0

    unlinked = 0
    deleted = 0
    for chunk in _chunks(reject_ids, WRITE_BATCH_SIZE):
        result = _run_batch(
            conn, lambda chunk=chunk: _bulk_reject(cur, chunk), "rejecting reviewed matches"
        )
        if result:
            unlinked += result[0]
            deleted += result[1]

    return confirmed, unlinked, deleted


def main():
    keep_ids, reject_ids, skipped, bad_rows = read_decisions(INPUT_PATH)
    print(f"Read {INPUT_PATH}: {len(keep_ids)} marked keep, {len(reject_ids)} marked reject, "
          f"{skipped} left blank.")

    if bad_rows:
        print(f"  {len(bad_rows)} row(s) had something unexpected and were skipped (only "
              f"'keep', 'reject', or blank are understood):")
        for raw_id, name, decision in bad_rows[:20]:
            print(f"    - {name!r} (id: {raw_id!r}, decision column said: {decision!r})")
        if len(bad_rows) > 20:
            print(f"    ...and {len(bad_rows) - 20} more")

    if not keep_ids and not reject_ids:
        print("Nothing to apply -- no rows were marked keep or reject.")
        return

    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            confirmed, unlinked, deleted = run_with_timeout(
                lambda: apply_all(conn, cur, keep_ids, reject_ids),
                DB_WRITE_HARD_TIMEOUT_SECONDS,
            )
        print(f"Done: {confirmed} product(s) confirmed as manual, {deleted} rejected product(s) "
              f"deleted ({unlinked} listing(s) unlinked from them).")
    except Exception as exc:  # noqa: BLE001 -- surface any failure plainly, then exit non-zero
        conn.rollback()
        print(f"ERROR applying reviewed matches: {type(exc).__name__}: {exc}", file=sys.stderr)
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
