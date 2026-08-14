"""
Exports every medium-confidence product match to a spreadsheet
(review_medium_matches.xlsx) so you can review all of them in one sitting,
in Excel or Google Sheets, instead of one at a time in the SQL editor.

Each row is one product, with every chain's listing that's currently linked
to it shown side by side, and a blank "decision" column with a dropdown:

    keep    -> the match is correct. Marks the product as manually
               confirmed (match_confidence = 'manual'), which also protects
               it from ever being automatically downgraded later.
    reject  -> the match is wrong. Unlinks every listing from that product
               and deletes the product row, so those listings go back to
               being unmatched and get a fresh look next time
               product_matcher.py runs.
    (blank) -> not reviewed yet. Left as medium confidence -- perfectly
               safe to leave as-is and come back to later.

See SETUP.md's "Reviewing medium-confidence matches in bulk" section for
the full step-by-step (download the file from this workflow's Artifacts,
fill in decisions, upload it back, run apply_reviewed_matches.py).

This only ever reads from the database -- it's always safe to re-run.
"""
import sys

import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from product_matcher import get_connection

OUTPUT_PATH = "review_medium_matches.xlsx"
FONT_NAME = "Arial"
HEADER_FILL = "4472C4"

QUERY = """
    SELECT product_id, canonical_name, size_value, size_unit,
           string_agg(brand || ': ' || chain_product_name, ' | ' ORDER BY brand)
               AS matched_listings
    FROM (
        SELECT DISTINCT product.id AS product_id, product.canonical_name,
               product.size_value, product.size_unit,
               store.brand, listing.chain_product_name
        FROM product
        JOIN listing ON listing.product_id = product.id
        JOIN outlet ON outlet.id = listing.outlet_id
        JOIN store ON store.id = outlet.store_id
        WHERE product.match_confidence = 'medium'
    ) sub
    GROUP BY product_id, canonical_name, size_value, size_unit
    ORDER BY product_id;
"""

INSTRUCTIONS = [
    ("How to review these matches", True, 13),
    ("", False, 11),
    ("Each row on the 'Review' tab is one product that the matcher linked across "
     "supermarkets with medium confidence -- probably correct, but not certain "
     "enough to trust automatically.", False, 11),
    ("", False, 11),
    ("For each row, compare the 'canonical_name' column to 'matched_listings'. If "
     "they're clearly the same product, click the decision cell and choose keep "
     "from the dropdown (or just type it). If they're clearly different products, "
     "choose reject.", False, 11),
    ("", False, 11),
    ("Leave decision blank for anything you're unsure about, or don't get to -- "
     "it's safe to leave as medium confidence and review it another time. Nothing "
     "changes in the database until you run the 'Apply reviewed matches' workflow "
     "with your filled-in file.", False, 11),
    ("", False, 11),
    ("Example -- same product, keep:", True, 11),
    ("  canonical_name:    Old Elpaso Bbq Fajita Kit 500g", False, 11),
    ("  matched_listings:  Greens: Old Elpaso Bbq Fajita Kit 500g | "
     "PAVI: OLD EL PASO BBQ FAJITA KI", False, 11),
    ("  decision:          keep", False, 11),
    ("", False, 11),
    ("Example -- different products, reject:", True, 11),
    ("  canonical_name:    Cauliflower Rice 400g", False, 11),
    ("  matched_listings:  Greens: Cauliflower Rice 400g | PAVI: CAULIFLOWER & BROC", False, 11),
    ("  decision:          reject", False, 11),
    ("", False, 11),
    ("Don't edit the product_id column. It's hidden by default (unhide it if you're "
     "curious) -- it's how the next step knows which row is which.", False, 11),
    ("", False, 11),
    ("When you're done (or done for now): save the file, upload it back into the "
     "repo with the exact same filename (review_medium_matches.xlsx), replacing "
     "the old one, then run the 'Apply reviewed matches' workflow from the Actions "
     "tab.", False, 11),
]


def fetch_rows(conn):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(QUERY)
        return cur.fetchall()


def _write_instructions_sheet(ws):
    ws.column_dimensions["A"].width = 100
    for i, (text, bold, size) in enumerate(INSTRUCTIONS, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=FONT_NAME, bold=bold, size=size)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 8 if not text else 18


def _write_review_sheet(ws, rows):
    headers = ["product_id", "canonical_name", "size", "matched_listings", "decision"]
    widths = [10, 36, 12, 75, 12]

    for col, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].hidden = True

    plain_font = Font(name=FONT_NAME)
    for r, row in enumerate(rows, start=2):
        size = ""
        if row["size_value"] is not None and row["size_unit"]:
            size = f"{row['size_value']} {row['size_unit']}"
        for col, value in enumerate(
            [str(row["product_id"]), row["canonical_name"], size, row["matched_listings"], ""],
            start=1,
        ):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = plain_font
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 4))

    last_row = max(len(rows) + 1, 2)
    dv = DataValidation(
        type="list",
        formula1='"keep,reject"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = "Please choose 'keep' or 'reject' from the dropdown, or leave it blank."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"E2:E{last_row}")


def build_workbook(rows):
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    _write_instructions_sheet(instructions)

    review = wb.create_sheet("Review")
    _write_review_sheet(review, rows)

    wb.active = review
    return wb


def main():
    conn = get_connection()
    try:
        rows = fetch_rows(conn)
    finally:
        conn.close()

    print(f"Found {len(rows)} medium-confidence product(s) to export.")
    if not rows:
        print("Nothing to export -- no medium-confidence matches right now.")

    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} ({len(rows)} row(s) plus instructions).")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001 -- surface any failure plainly, then exit non-zero
        print(f"ERROR during export: {type(exc).__name__}: {exc}", file=sys.stderr)
        sys.exit(1)
